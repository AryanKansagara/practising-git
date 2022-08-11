from openpyxl import workbook, load_workbook

wb = load_workbook('data.xlsx')

ws = wb.active

# print(ws['A2'].value) 

#this can change the value at A2
# ws['A2'].value = 66



#this creates a sheet
# wb.create_sheet("name_of_the_sheet")



#this shows the names of the sheets
# print(wb.sheetnames)



#save/update the workbook
wb.save('data.xlsx')
